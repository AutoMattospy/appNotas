from playwright.sync_api import Playwright, sync_playwright, expect
import time
import datetime
from openpyxl import load_workbook , workbook
import openpyxl
import os
import re

def check_text_on_screen(page1):
    try:
        locator = page1.locator('section.estrutura_componente_modal.contexto_acessibilidade.ui-draggable')
        if locator.is_visible():
            #print(f'Texto esta visivel na tela.')
            return True
        else:
            #print(f'texto não esta visivel na tela.')
            return False
    except Exception as e:
        print(f'Ocorreu um erro: {e}')
        return False

def tratarCNPJCPF(variable):
    #Substitui todos os caracteres que não são números por uma string vazia
    return re.sub(r'\D', '', variable)


extension_path1 = "C:\\python\\ÁreadeTrabalho\\nativa\\extension\\reCAPTCHA"
extension_path2 ="C:\\python\\ÁreadeTrabalho\\nativa\\extension\\reCAPTCHA2"
user_data_dir ="C:\\python\\ÁreadeTrabalho\\nativa\\extension"
def run(playwright: Playwright) -> None:
    
    browser_context = playwright.chromium.launch_persistent_context(
        user_data_dir=user_data_dir,
        channel="chrome", 
        headless=False,
        args=[
            f'--disable-extensions-except={extension_path1},{extension_path2}',
            f'--load-extension={extension_path1},{extension_path2}'
        ]
        )

    page = browser_context.new_page()

    if len(browser_context.pages) > 0:
        browser_context.pages[0].close()

    page.goto("https://nfse-doisirmaos.atende.net/autoatendimento/servicos/nfse?redirected=1")
    page.get_by_placeholder("CPF ou CNPJ").click() 
    page.get_by_placeholder("CPF ou CNPJ").fill("")
    page.get_by_placeholder("Senha").click()
    page.get_by_placeholder("Senha").fill("")
    page.get_by_role("button", name="Entrar", exact=True).click()
    page.get_by_role("link", name="Acessar").click()
    time.sleep(3)
    #page.reload()
    #page.frame_locator("iframe[name=\"a-b0lkgxhwcgia\"]").get_by_label("Não sou um robô").click()
    with page.expect_popup(timeout=60000) as page1_info:
        # Aguarda até que o elemento de verificação de acesso não seja visível
        print("Aguarde enquanto o CAPTCHA é resolvido...")
        while True:
            if not page.query_selector("role=heading[name=\"Verificação de acesso\"]"):
                print("Elemento não encontrado! O CAPTCHA foi resolvido.")
                time.sleep(3) # espera 3 segundos e continua o processo
                break
            else:
                print("Elemento encontrado. O CAPTCHA ainda não foi resolvido. Tentando novamente...")
                time.sleep(5)  # espera 5 segundos antes de tentar novamente
        page1 = page1_info.value
    
    # conecta na planilha
    workbook = openpyxl.load_workbook()  #Tem que estar na mesma pasta
    
    planilha = workbook['Planilha1']
    #planilha = workbook.active
    
    # Abre a aba.
    sheet = workbook.get_sheet_by_name('Planilha1')

    planilha = sheet
    # Obtenha os valores da coluna A.
    for row in range(2, planilha.max_row + 1):

        #Coluna Matr. - Nome do Aluno
        celula = planilha.cell(row=row, column=1)
        MatrNomedoAluno1 = celula.value
        #TRATAR NOMES
        MatrNomedoAluno1 = re.sub(r'[^a-zA-Z\s]', '', MatrNomedoAluno1)



        #COLUNA CNPJ/CPF
        celula = planilha.cell(row=row, column=2)
        CNPJ_CPF = celula.value

        #tratando dos dados
        CNPJ_CPF = str(CNPJ_CPF)
        CNPJ_CPF = ''.join(filter(str.isdigit, CNPJ_CPF))

        if len(CNPJ_CPF) < 11:
            while len(CNPJ_CPF) < 11:
                CNPJ_CPF = CNPJ_CPF.zfill(11)

        #print(CNPJ_CPF)
        
        #Coluna Mes
        celula = planilha.cell(row=row, column=3)
        ValorMes = celula.value
        #tratamwnro de dados
        ValorMes = str(ValorMes)
        ValorMes = ValorMes.replace(".",",")
        

        #coluna Status
        celula = planilha.cell(row=row, column=4)
        status = celula.value
            
        #Variaveis:
        cpf_cliente = list(CNPJ_CPF)
        dataAtual= datetime.datetime.now()

        #Executa a emição da nota se o status for diferente de "ok"
        if status != "ok":
            if status != "Erro! Por favor verifique o cadastro":

                try:                
                    page1.wait_for_timeout(5000)
                    page1.get_by_label("Conjuntos").locator("div").filter(has_text="Visão Gerencial").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Visões Gerenciais Disponíveis").locator("div").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_text("Emitir Nota Fiscal").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Próximo").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_placeholder("Pesquisar por CPF").click()
                    page1.wait_for_timeout(1000)

                    #Digita o CNPJ
                    page1.get_by_placeholder("Pesquisar por CPF").click()
                    for char in cpf_cliente:
                        page1.keyboard.type(char)
                        page1.wait_for_timeout(1000)
                    
                    #Verifica o erro:
                    is_visible = check_text_on_screen(page1)
                    if is_visible:
                        page1.wait_for_timeout(5000)
                        page1.get_by_role("button", name="Sim").click()
                        page1.wait_for_timeout(1000)
                        page1.get_by_placeholder("000.000.000-").click()
                        #
                        clickVoltar = 12
                        for i in  range(clickVoltar):
                            page1.get_by_placeholder("000.000.000-").press("ArrowLeft")
                            #print(f'volta:{i + 1}')

                        page1.wait_for_timeout(1000)
                        page1.keyboard.type(CNPJ_CPF)
                        page1.wait_for_timeout(1000)
                        page1.get_by_label("Nome Completo").click()
                        page1.wait_for_timeout(1000)
                        page1.get_by_label("Nome Completo").fill(MatrNomedoAluno1)
                        page1.wait_for_timeout(1000)
                        page1.get_by_label("Aba - Endereço Principal").locator("input[name=\"LogradouroBairro\\.Bairro\\.nome\"]").click()
                        page1.wait_for_timeout(1000)
                        page1.keyboard.type("Primavera")
                        page1.wait_for_timeout(1000)
                        #page1.get_by_label("Aba - Endereço Principal").locator("input[name=\"LogradouroBairro\\.Bairro\\.nome\"]").fill("Primavera")    
                        page1.get_by_role("cell", name="Dois Irmãos").click()
                        page1.wait_for_timeout(1000)
                        page1.get_by_label("Aba - Endereço Principal").locator("input[name=\"LogradouroBairro\\.Logradouro\\.nome\"]").click()
                        page1.wait_for_timeout(1000)
                        page1.keyboard.type("SAO LEOPOLDO")
                        page1.get_by_label("Aba - Endereço Principal").locator("input[name=\"LogradouroBairro\\.Logradouro\\.nome\"]").fill()
                        page1.wait_for_timeout(1000)
                        page1.get_by_text("SAO").first.click()
                        page1.wait_for_timeout(1000)
                        page1.get_by_label("Janela - Incluir Pessoa Física").get_by_role("button", name="Próximo").click()
                        page1.wait_for_timeout(1000)
                        page1.locator("label").filter(has_text="Não Informado").locator("span").first.click()
                        page1.wait_for_timeout(1000)
                        page1.get_by_role("button", name="Confirmar").click()



                    page1.wait_for_timeout(3000)
                    page1.get_by_placeholder("Pesquisar por CPF").press("Enter")
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Próximo").click()
                    page1.wait_for_timeout(2000)
                    page1.get_by_label("Local da Prestação").fill("8625")
                    page1.wait_for_timeout(2000)
                    page1.get_by_role("textbox", name="Digite aqui para consultar").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Valor do Serviço").click() 
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Valor do Serviço").fill(ValorMes)
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Descrição").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Descrição").press("CapsLock")
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Descrição").fill("ACADEMIA")
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Lista de Serviço").select_option("604")
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Próximo").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_text("Inf. Complementares").click()
                    page1.wait_for_timeout(1000)
                    page1.get_by_label("Imprimir após confirmação?").uncheck()
                    page1.wait_for_timeout(1000)
                    page1.get_by_role("button", name="Emitir").click()
                    page1.wait_for_timeout(2000)                

                    #Da um "ok"
                    celula = planilha.cell(row=row, column=4)
                    celula.value = 'ok'
                    
                    

                    workbook.save() #Salva a planilha


                except:
                    #Indica erro na planilha
                    celula = planilha.cell(row=row, column=4)
                    celula.value = 'Erro! Por favor verifique o cadastro'
                    
                    workbook.save() #Salva a planilha
        

    # ---------------------
    #context.close()
    browser_context.close()

    
    # Substitua 'caminho/para/seu/arquivo.xlsx' pelo caminho do seu arquivo Excel
    file_path = "C:\\python\\ÁreadeTrabalho\\nativa\\EmisordeNotas.xlsm"

    # Abrir o arquivo no programa padrão do sistema operacional
    os.startfile(file_path)

with sync_playwright() as playwright:
    run(playwright)
